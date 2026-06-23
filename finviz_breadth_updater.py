"""Update Finviz market breadth data files.

Fetches the Finviz homepage, extracts the four requested breadth counts and
percentages, and updates CSV/XLSX outputs using one row per New York market date.
"""

from __future__ import annotations

import csv
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

FINVIZ_URL = "https://finviz.com/"
OUTPUT_COLUMNS = [
    "Date",
    "New High",
    "New High %",
    "New Low",
    "New Low %",
    "Advancing",
    "Advancing %",
    "Declining",
    "Declining %",
]
LEGACY_OUTPUT_COLUMNS = ["Date", "New High", "New Low", "Advancing", "Declining"]
VALUE_LABELS = ["New High", "New Low", "Advancing", "Declining"]
CSV_PATH = Path("finviz_breadth.csv")
XLSX_PATH = Path("finviz_breadth.xlsx")
LOG_PATH = Path("finviz_breadth_log.txt")
NY_TZ = ZoneInfo("America/New_York")
TELEGRAM_API_BASE = "https://api.telegram.org"


class FinvizBreadthError(RuntimeError):
    """Raised when Finviz data cannot be fetched or parsed clearly."""


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(LOG_PATH, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def fetch_finviz_homepage() -> str:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/126.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
    }

    try:
        response = requests.get(FINVIZ_URL, headers=headers, timeout=30)
    except requests.RequestException as exc:
        raise FinvizBreadthError(f"Finviz request failed: {exc}") from exc

    if response.status_code in {401, 403, 407, 429, 503}:
        raise FinvizBreadthError(
            f"Finviz may have blocked the request; HTTP status {response.status_code}."
        )
    if not response.ok:
        raise FinvizBreadthError(f"Finviz request failed with HTTP status {response.status_code}.")

    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    visible_text = soup.get_text(" ", strip=True).lower()
    title_text = soup.title.get_text(" ", strip=True).lower() if soup.title else ""
    block_check_text = f"{title_text} {visible_text[:2_000]}"
    block_markers = [
        "access denied",
        "captcha",
        "verify you are human",
        "unusual traffic",
    ]
    if any(marker in block_check_text for marker in block_markers):
        raise FinvizBreadthError("Finviz returned a blocking or verification page.")
    if len(html) < 10_000:
        raise FinvizBreadthError("Finviz response was unexpectedly short; page may be blocked.")

    return html


def normalized_page_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ", strip=True)
    return re.sub(r"\s+", " ", text)


def segment_after_label(page_text: str, label: str) -> str:
    label_match = re.search(rf"\b{re.escape(label)}\b", page_text, re.IGNORECASE)
    if not label_match:
        raise FinvizBreadthError(f"Could not find '{label}'. Finviz layout may have changed.")

    segment = page_text[label_match.end() : label_match.end() + 180]
    stop_markers = VALUE_LABELS + ["Above", "Below", "SMA50", "SMA200"]
    marker_pattern = re.compile(
        r"\b(?:" + "|".join(re.escape(marker) for marker in stop_markers if marker != label) + r")\b",
        re.IGNORECASE,
    )
    next_marker = marker_pattern.search(segment)
    if next_marker:
        segment = segment[: next_marker.start()]

    return segment


def extract_breadth_value(page_text: str, label: str) -> Dict[str, str]:
    segment = segment_after_label(page_text, label)
    count_match = re.search(r"\(\s*(?P<count>[\d,]+)\s*\)", segment)
    percent_match = re.search(r"(?P<percent>\d+(?:\.\d+)?)\s*%", segment)

    if not count_match:
        raise FinvizBreadthError(
            f"Could not find parenthesized count for '{label}'. Finviz layout may have changed."
        )
    if not percent_match:
        raise FinvizBreadthError(
            f"Could not find percentage for '{label}'. Finviz layout may have changed."
        )

    count_text = count_match.group("count").replace(",", "")
    try:
        count = int(count_text)
    except ValueError as exc:
        raise FinvizBreadthError(f"Invalid count for '{label}': {count_match.group('count')}") from exc

    return {"count": str(count), "percent": f"{percent_match.group('percent')}%"}


def extract_breadth_values(html: str) -> Dict[str, Dict[str, str]]:
    page_text = normalized_page_text(html)
    values = {label: extract_breadth_value(page_text, label) for label in VALUE_LABELS}

    missing = [label for label in VALUE_LABELS if label not in values]
    if missing:
        raise FinvizBreadthError(f"Missing breadth values: {', '.join(missing)}")

    return values


def read_existing_rows() -> List[Dict[str, str]]:
    if not CSV_PATH.exists():
        return []

    with CSV_PATH.open("r", newline="", encoding="utf-8") as csv_file:
        reader = csv.DictReader(csv_file)
        if reader.fieldnames == LEGACY_OUTPUT_COLUMNS:
            migrated_rows = []
            for row in reader:
                migrated_row = {column: "" for column in OUTPUT_COLUMNS}
                for column in LEGACY_OUTPUT_COLUMNS:
                    migrated_row[column] = row.get(column, "")
                migrated_rows.append(migrated_row)
            return migrated_rows
        if reader.fieldnames != OUTPUT_COLUMNS:
            raise FinvizBreadthError(
                f"Existing CSV columns are {reader.fieldnames}; expected {OUTPUT_COLUMNS}."
            )
        return [{column: row.get(column, "") for column in OUTPUT_COLUMNS} for row in reader]


def write_csv(rows: List[Dict[str, str]]) -> None:
    with CSV_PATH.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: List[Dict[str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Finviz Breadth"

    worksheet.append(OUTPUT_COLUMNS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row[column] for column in OUTPUT_COLUMNS])

    for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        width = max(len(column_name), 12)
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = width

    workbook.save(XLSX_PATH)


def upsert_today_row(values: Dict[str, Dict[str, str]]) -> List[Dict[str, str]]:
    market_date = current_market_date()
    rows = read_existing_rows()
    today_row = {"Date": market_date}
    for label in VALUE_LABELS:
        today_row[label] = values[label]["count"]
        today_row[f"{label} %"] = values[label]["percent"]

    updated = False
    for index, row in enumerate(rows):
        if row["Date"] == market_date:
            rows[index] = today_row
            updated = True
            break

    if not updated:
        rows.append(today_row)

    rows.sort(key=lambda item: item["Date"])
    return rows


def current_market_date() -> str:
    return datetime.now(NY_TZ).date().isoformat()


def build_telegram_message(market_date: str, values: Dict[str, Dict[str, str]]) -> str:
    return (
        f"Finviz Breadth — {market_date}\n\n"
        f"New High: {values['New High']['count']} ({values['New High']['percent']})\n"
        f"New Low: {values['New Low']['count']} ({values['New Low']['percent']})\n"
        f"Advancing: {values['Advancing']['count']} ({values['Advancing']['percent']})\n"
        f"Declining: {values['Declining']['count']} ({values['Declining']['percent']})"
    )


def send_telegram_message(market_date: str, values: Dict[str, Dict[str, str]]) -> None:
    bot_token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")

    if not bot_token or not chat_id:
        logging.info("Telegram delivery skipped because secrets are not configured.")
        return

    message = build_telegram_message(market_date, values)
    url = f"{TELEGRAM_API_BASE}/bot{bot_token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": message,
        "disable_web_page_preview": True,
    }

    try:
        response = requests.post(url, data=payload, timeout=30)
    except requests.RequestException:
        raise FinvizBreadthError("Telegram request failed.") from None

    if not response.ok:
        raise FinvizBreadthError(
            f"Telegram delivery failed with HTTP status {response.status_code}."
        )

    logging.info("Telegram delivery completed successfully.")


def main() -> None:
    configure_logging()
    logging.info("Starting Finviz breadth update.")

    try:
        html = fetch_finviz_homepage()
        values = extract_breadth_values(html)
        logging.info("Extracted breadth values: %s", values)

        rows = upsert_today_row(values)
        write_csv(rows)
        write_xlsx(rows)

        logging.info("Updated %s and %s with columns: %s", CSV_PATH, XLSX_PATH, OUTPUT_COLUMNS)
        send_telegram_message(current_market_date(), values)
        logging.info("Finviz breadth update completed successfully.")
    except Exception as exc:
        logging.exception("Finviz breadth update failed clearly: %s", exc)
        raise


if __name__ == "__main__":
    main()

